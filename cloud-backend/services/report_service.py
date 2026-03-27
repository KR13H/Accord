from __future__ import annotations

import io
from decimal import Decimal
from datetime import date
from typing import Any, Callable

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


class ReportService:
    """Financial report engine for Balance Sheet, P&L, and ratio analysis."""

    def __init__(
        self,
        *,
        money: Callable[[Any], Decimal],
        money_str: Callable[[Any], str],
    ) -> None:
        self.money = money
        self.money_str = money_str

    def build_balance_sheet(self, accounts: list[dict[str, Any]]) -> dict[str, Any]:
        assets = Decimal("0")
        liabilities = Decimal("0")
        equity = Decimal("0")

        lines: list[dict[str, Any]] = []
        for row in accounts:
            name = str(row.get("name") or "")
            acc_type = str(row.get("type") or "")
            amount = self.money(str(row.get("balance") or "0"))
            lines.append({"name": name, "type": acc_type, "balance": self.money_str(amount)})
            if acc_type == "Asset":
                assets += amount
            elif acc_type == "Liability":
                liabilities += amount
            elif acc_type == "Equity":
                equity += amount

        return {
            "status": "ok",
            "report": "BALANCE_SHEET",
            "totals": {
                "assets": self.money_str(assets),
                "liabilities": self.money_str(liabilities),
                "equity": self.money_str(equity),
                "liabilities_plus_equity": self.money_str(liabilities + equity),
            },
            "lines": lines,
        }

    def build_profit_and_loss(self, accounts: list[dict[str, Any]]) -> dict[str, Any]:
        revenue = Decimal("0")
        expenses = Decimal("0")

        for row in accounts:
            acc_type = str(row.get("type") or "")
            amount = self.money(str(row.get("balance") or "0"))
            if acc_type == "Revenue":
                revenue += amount
            elif acc_type == "Expense":
                expenses += amount

        net_profit = revenue - expenses
        return {
            "status": "ok",
            "report": "PROFIT_AND_LOSS",
            "summary": {
                "revenue": self.money_str(revenue),
                "expenses": self.money_str(expenses),
                "net_profit": self.money_str(net_profit),
                "margin_pct": round((float(net_profit / revenue) * 100.0), 2) if revenue != 0 else 0.0,
            },
        }

    def build_ratio_analysis(self, accounts: list[dict[str, Any]]) -> dict[str, Any]:
        balance = self.build_balance_sheet(accounts)
        pnl = self.build_profit_and_loss(accounts)

        assets = self.money(balance["totals"]["assets"])
        liabilities = self.money(balance["totals"]["liabilities"])
        equity = self.money(balance["totals"]["equity"])
        current_ratio = float(assets / liabilities) if liabilities != 0 else 0.0
        debt_to_equity = float(liabilities / equity) if equity != 0 else 0.0

        return {
            "status": "ok",
            "report": "RATIO_ANALYSIS",
            "ratios": {
                "current_ratio": round(current_ratio, 4),
                "debt_to_equity": round(debt_to_equity, 4),
                "net_margin_pct": pnl["summary"]["margin_pct"],
            },
        }

    def build_ca_monthly_payload(
        self,
        *,
        period_start: date,
        period_end: date,
        ledger_entries: list[dict[str, Any]],
        heatmap: dict[str, Any],
        friday_summary: dict[str, Any],
        market_context: dict[str, Any],
    ) -> dict[str, Any]:
        total_entries = len(ledger_entries)
        total_debit = Decimal("0")
        total_credit = Decimal("0")
        for row in ledger_entries:
            total_debit += self.money(row.get("total_debit") or "0")
            total_credit += self.money(row.get("total_credit") or "0")

        unresolved_blockers = int(friday_summary.get("ims_actionables", {}).get("PENDING", 0))
        reversal = friday_summary.get("reversal_risks", {})
        vendor_ranking = friday_summary.get("vendor_risk_ranking", [])

        return {
            "status": "ok",
            "report": "CA_MONTHLY_RISK_AND_COMPLIANCE",
            "period": {
                "from": period_start.isoformat(),
                "to": period_end.isoformat(),
            },
            "kpis": {
                "ledger_entries": total_entries,
                "ledger_total_debit": self.money_str(total_debit),
                "ledger_total_credit": self.money_str(total_credit),
                "unresolved_gst_blockers": unresolved_blockers,
                "open_alerts_total": int(heatmap.get("open_alerts_total", 0)),
                "aggregate_risk": str(heatmap.get("aggregate_risk") or "LOW"),
                "market_risk_level": str(market_context.get("risk_level") or "MEDIUM"),
                "itc_at_risk": str(friday_summary.get("summary", {}).get("total_itc_at_risk") or "0.0000"),
            },
            "reversal_risk": {
                "immediate_reversal_risk": str(reversal.get("immediate_reversal_risk") or "0.0000"),
                "projected_annual_interest_18pct": str(reversal.get("projected_annual_interest_18pct") or "0.0000"),
                "at_risk_invoice_count": int(reversal.get("at_risk_invoice_count", 0)),
                "at_risk_references": reversal.get("at_risk_references", []),
                "status": str(reversal.get("status") or "MONITOR"),
            },
            "market_intel": {
                "risk_level": str(market_context.get("risk_level") or "MEDIUM"),
                "source_kind": str(market_context.get("source_kind") or "UNKNOWN"),
                "created_at": str(market_context.get("created_at") or ""),
                "trend_summary": str(market_context.get("trend_summary") or "No market summary available"),
            },
            "top_vendor_risks": vendor_ranking[:10],
            "heatmap": {
                "risk_buckets": heatmap.get("risk_buckets", {}),
                "open_alerts_total": int(heatmap.get("open_alerts_total", 0)),
                "cells": heatmap.get("cells", [])[:50],
            },
            "ledger_entries": ledger_entries[:200],
        }

    def generate_ca_monthly_pdf(self, payload: dict[str, Any]) -> bytes:
        buffer = io.BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 52

        def line(text: str, gap: int = 16, font: str = "Helvetica", size: int = 10) -> None:
            nonlocal y
            if y <= 48:
                pdf.showPage()
                y = height - 52
            pdf.setFont(font, size)
            pdf.drawString(42, y, text)
            y -= gap

        period = payload.get("period", {})
        kpis = payload.get("kpis", {})
        reversal = payload.get("reversal_risk", {})
        market = payload.get("market_intel", {})
        risk_buckets = payload.get("heatmap", {}).get("risk_buckets", {})

        pdf.setTitle("Accord Monthly Risk & Compliance Report")
        line("ACCORD - MONTHLY RISK & COMPLIANCE REPORT", font="Helvetica-Bold", size=14, gap=22)
        line(f"Period: {period.get('from', '')} to {period.get('to', '')}", font="Helvetica-Bold", size=11)
        line("", gap=8)

        line("Key Performance Indicators", font="Helvetica-Bold", size=11, gap=18)
        line(f"Ledger Entries: {kpis.get('ledger_entries', 0)}")
        line(f"Ledger Total Debit: INR {kpis.get('ledger_total_debit', '0.0000')}")
        line(f"Ledger Total Credit: INR {kpis.get('ledger_total_credit', '0.0000')}")
        line(f"Unresolved GST Blockers: {kpis.get('unresolved_gst_blockers', 0)}")
        line(f"Open Alerts: {kpis.get('open_alerts_total', 0)}")
        line(f"Aggregate Risk: {kpis.get('aggregate_risk', 'LOW')}")
        line(f"Market Risk: {kpis.get('market_risk_level', 'MEDIUM')}")
        line("", gap=8)

        line("Rule 37A Reversal Risk", font="Helvetica-Bold", size=11, gap=18)
        line(f"Immediate Reversal Risk: INR {reversal.get('immediate_reversal_risk', '0.0000')}")
        line(f"Projected Interest @18%: INR {reversal.get('projected_annual_interest_18pct', '0.0000')}")
        line(f"At-Risk Invoice Count: {reversal.get('at_risk_invoice_count', 0)}")
        line(f"Status: {reversal.get('status', 'MONITOR')}")
        line("", gap=8)

        line("Market Intelligence", font="Helvetica-Bold", size=11, gap=18)
        line(f"Source: {market.get('source_kind', 'UNKNOWN')}")
        line(f"Risk Level: {market.get('risk_level', 'MEDIUM')}")
        line(f"Updated At: {market.get('created_at', '')}")
        line(f"Trend Summary: {market.get('trend_summary', '')}")
        line("", gap=8)

        line("Heatmap Risk Buckets", font="Helvetica-Bold", size=11, gap=18)
        line(f"LOW: {risk_buckets.get('LOW', 0)}")
        line(f"MEDIUM: {risk_buckets.get('MEDIUM', 0)}")
        line(f"HIGH: {risk_buckets.get('HIGH', 0)}")
        line(f"CRITICAL: {risk_buckets.get('CRITICAL', 0)}")

        vendors = payload.get("top_vendor_risks", [])
        if vendors:
            line("", gap=8)
            line("Top Vendor Risks", font="Helvetica-Bold", size=11, gap=18)
            for idx, vendor in enumerate(vendors[:10], start=1):
                gstin = str(vendor.get("gstin") or "")
                legal_name = str(vendor.get("legal_name") or "Unknown Vendor")
                score = str(vendor.get("filing_consistency_score") or "0")
                itc = str(vendor.get("total_itc_at_risk") or "0.0000")
                line(f"{idx}. {legal_name} ({gstin}) | Score {score} | ITC Risk INR {itc}")

        pdf.showPage()
        pdf.save()
        content = buffer.getvalue()
        buffer.close()
        return content

    def generate_ca_monthly_excel(self, payload: dict[str, Any]) -> bytes:
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        period = payload.get("period", {})
        kpis = payload.get("kpis", {})
        reversal = payload.get("reversal_risk", {})
        market = payload.get("market_intel", {})

        ws_summary.append(["Accord Monthly Risk & Compliance Report"])
        ws_summary.append(["Period From", period.get("from", "")])
        ws_summary.append(["Period To", period.get("to", "")])
        ws_summary.append([])
        ws_summary.append(["KPI", "Value"])
        for key, value in kpis.items():
            ws_summary.append([key, value])

        ws_summary.append([])
        ws_summary.append(["Reversal Risk", "Value"])
        for key, value in reversal.items():
            if isinstance(value, list):
                ws_summary.append([key, ", ".join([str(item) for item in value[:15]])])
            else:
                ws_summary.append([key, value])

        ws_summary.append([])
        ws_summary.append(["Market Intel", "Value"])
        for key, value in market.items():
            ws_summary.append([key, value])

        ws_vendors = wb.create_sheet("Vendor Risks")
        ws_vendors.append([
            "GSTIN",
            "Legal Name",
            "Filing Score",
            "Avg Delay Days",
            "Last Filed",
            "ITC At Risk",
            "Payment Advice",
        ])
        for row in payload.get("top_vendor_risks", []):
            ws_vendors.append(
                [
                    row.get("gstin", ""),
                    row.get("legal_name", ""),
                    row.get("filing_consistency_score", ""),
                    row.get("avg_filing_delay_days", ""),
                    row.get("last_gstr1_filed_at", ""),
                    row.get("total_itc_at_risk", ""),
                    row.get("payment_advice", ""),
                ]
            )

        ws_ledger = wb.create_sheet("Ledger Entries")
        ws_ledger.append(["Entry ID", "Date", "Reference", "Description", "Total Debit", "Total Credit"])
        for row in payload.get("ledger_entries", []):
            ws_ledger.append(
                [
                    row.get("id", ""),
                    row.get("date", ""),
                    row.get("reference", ""),
                    row.get("description", ""),
                    row.get("total_debit", ""),
                    row.get("total_credit", ""),
                ]
            )

        ws_heatmap = wb.create_sheet("Heatmap")
        ws_heatmap.append([
            "GSTIN",
            "Vendor",
            "Risk Level",
            "Trust Score",
            "ITC At Risk",
            "Open Alerts",
            "Payment Advice",
        ])
        for row in payload.get("heatmap", {}).get("cells", []):
            ws_heatmap.append(
                [
                    row.get("gstin", ""),
                    row.get("vendor_name", ""),
                    row.get("risk_level", ""),
                    row.get("trust_score", ""),
                    row.get("total_itc_at_risk", ""),
                    row.get("open_alert_count", ""),
                    row.get("payment_advice", ""),
                ]
            )

        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        output.close()
        return content
